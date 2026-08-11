import streamlit as st
import openpyxl
import io
import zipfile

st.set_page_config(page_title="Faturamento EBM Quintto", layout="centered")

st.title("📊 Gerador de Faturamento - EBM Quintto")
st.write("Preencha os dados da Nota Fiscal abaixo para gerar as 3 planilhas de faturamento automaticamente.")

# Função para escrever em células mescladas de forma segura
def safe_write(ws, cell_address, value):
    cell = ws[cell_address]
    if type(cell).__name__ == 'MergedCell':
        for merged_range in ws.merged_cells.ranges:
            if cell_address in merged_range:
                ws.cell(row=merged_range.min_row, column=merged_range.min_col, value=value)
                return
    else:
        ws[cell_address] = value

with st.form("faturamento_form"):
    col1, col2 = st.columns(2)
    tipo = col1.selectbox("Tipo de Serviço", ["Mídia", "Produção", "Custos Internos"])
    empenho = col2.text_input("Nº do Empenho")
    
    nup = st.text_input("NUP (Ex: 30001.008502/2026-87)")
    descricao = st.text_area("Descrição do Serviço / PI / Plano")
    
    col3, col4 = st.columns(2)
    nf_ebm = col3.text_input("Nº NF EBM")
    data_ebm = col4.text_input("Data Emissão EBM (DD/MM/AAAA)")
    
    fornecedor = st.text_input("Razão Social do Fornecedor / Veículo")
    cnpj_forn = st.text_input("CNPJ do Fornecedor")
    
    col5, col6 = st.columns(2)
    nf_forn = col5.text_input("Nº NF Fornecedor")
    data_forn = col6.text_input("Data Emissão Fornecedor (DD/MM/AAAA)")
    
    col7, col8 = st.columns(2)
    valor_bruto = col7.number_input("Valor Bruto Total da NF (R$)", min_value=0.0, step=0.01)
    valor_fornecedor = col8.number_input("Valor Repassado ao Fornecedor (R$)", min_value=0.0, step=0.01)
    
    simples_nacional = st.radio("O Fornecedor é optante do Simples Nacional?", ["SIM", "NÃO"])
    
    submit = st.form_submit_button("Gerar Documentos")

if submit:
    try:
        # Cálculos Financeiros
        honorarios_ebm = valor_bruto - valor_fornecedor
        
        if simples_nacional == "SIM":
            ret_iss_forn = 0.0
            ret_irrf_forn = 0.0
        else:
            ret_iss_forn = valor_fornecedor * 0.02
            ret_irrf_forn = valor_fornecedor * 0.048
            
        ret_imp_forn = ret_iss_forn + ret_irrf_forn
        valor_liq_forn = valor_fornecedor - ret_imp_forn
        
        ret_iss_ebm = honorarios_ebm * 0.05
        ret_irrf_ebm = honorarios_ebm * 0.048
        ret_imp_ebm = ret_iss_ebm + ret_irrf_ebm
        valor_liq_ebm = honorarios_ebm - ret_imp_ebm
        
        valor_liq_nf = valor_liq_forn + valor_liq_ebm

        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w") as zf:
            
            # 1. PLANILHA CARTA
            wb_carta = openpyxl.load_workbook("01_CARTA.xlsx")
            sheet_carta = wb_carta.active
            safe_write(sheet_carta, "B1", empenho)
            safe_write(sheet_carta, "B2", nup)
            safe_write(sheet_carta, "B3", nf_ebm)
            safe_write(sheet_carta, "B4", data_ebm)
            safe_write(sheet_carta, "B5", valor_bruto)
            safe_write(sheet_carta, "B6", descricao)
            
            excel_carta = io.BytesIO()
            wb_carta.save(excel_carta)
            zf.writestr(f"01_CARTA_NF{nf_ebm}.xlsx", excel_carta.getvalue())

            # 2. PLANILHA FINANCEIRA
            wb_fin = openpyxl.load_workbook("02_PLANILHA FINANCEIRA.xlsx")
            sheet_fin = wb_fin.active
            safe_write(sheet_fin, "B1", empenho)
            safe_write(sheet_fin, "B2", nup)
            safe_write(sheet_fin, "B3", tipo)
            safe_write(sheet_fin, "B4", fornecedor)
            safe_write(sheet_fin, "B5", cnpj_forn)
            safe_write(sheet_fin, "B6", nf_forn)
            safe_write(sheet_fin, "B7", valor_bruto)
            safe_write(sheet_fin, "B8", valor_fornecedor)
            safe_write(sheet_fin, "B9", honorarios_ebm)
            
            excel_fin = io.BytesIO()
            wb_fin.save(excel_fin)
            zf.writestr(f"02_PLANILHA_FINANCEIRA_NF{nf_ebm}.xlsx", excel_fin.getvalue())

            # 3. DADOS PARA LIQUIDAR
            wb_liq = openpyxl.load_workbook("17_PLANILHA DADOS PARA LIQUIDAR.xlsx")
            sheet_liq = wb_liq.active
            safe_write(sheet_liq, "B1", empenho)
            safe_write(sheet_liq, "B2", nup)
            safe_write(sheet_liq, "B3", valor_bruto)
            safe_write(sheet_liq, "B4", (ret_iss_ebm + ret_iss_forn))
            safe_write(sheet_liq, "B5", (ret_irrf_ebm + ret_irrf_forn))
            safe_write(sheet_liq, "B6", valor_liq_nf)
            
            safe_write(sheet_liq, "A9", nf_ebm)
            safe_write(sheet_liq, "B9", data_ebm)
            safe_write(sheet_liq, "E9", ret_iss_ebm)
            safe_write(sheet_liq, "F9", ret_irrf_ebm)
            safe_write(sheet_liq, "G9", honorarios_ebm)
            
            safe_write(sheet_liq, "A13", nf_forn)
            safe_write(sheet_liq, "B13", data_forn)
            safe_write(sheet_liq, "C13", cnpj_forn)
            safe_write(sheet_liq, "D13", fornecedor)
            safe_write(sheet_liq, "E13", ret_iss_forn)
            safe_write(sheet_liq, "F13", ret_irrf_forn)
            safe_write(sheet_liq, "G13", valor_fornecedor)
            safe_write(sheet_liq, "H13", simples_nacional)
            
            excel_liq = io.BytesIO()
            wb_liq.save(excel_liq)
            zf.writestr(f"03_DADOS_LIQUIDAR_NF{nf_ebm}.xlsx", excel_liq.getvalue())
            
        st.success("Tudo pronto! As 3 planilhas foram geradas com sucesso.")
        st.download_button(label="📥 Baixar Faturamento (ZIP)", data=zip_buffer.getvalue(), file_name=f"Faturamento_NF{nf_ebm}.zip", mime="application/zip")

    except Exception as e:
        st.error(f"Ocorreu um erro ao gerar os documentos: {e}")
